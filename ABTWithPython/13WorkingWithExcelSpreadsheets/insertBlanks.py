# -*- coding: utf-8 -*-
"""
Created on Sat Oct 24 23:07:29 2020

@author: Ankit Parashar
"""
#! python3
# blankRowInserter.py
import openpyxl, sys

N = int(sys.argv[1])
M = int(sys.argv[2])
fileName = sys.argv[3]

wb = openpyxl.load_workbook(fileName)
sheet = wb.active
max_cols = sheet.max_column
max_rows = sheet.max_row

wb1 = openpyxl.Workbook()
sheet1 = wb1.active
for i in range(1, max_rows+1):
    for j in range(1, max_cols+1):
        if i >= N:
            sheet1.cell(row = i+M, column=j).value = sheet.cell(row = i, column = j).value
        else:
            sheet1.cell(row = i, column=j).value = sheet.cell(row = i, column = j).value
    
wb1.save(fileName+'_edited.xlsx')