# -*- coding: utf-8 -*-
"""
Created on Sat Oct 24 14:23:24 2020

@author: Ankit Parashar
"""
#! python3
# multiplicationTable.py

import sys, openpyxl
from openpyxl.styles import Font

n = int(sys.argv[1])
wb = openpyxl.Workbook()
sheet = wb.active
fontObj = Font(bold=True)
for i in range(2, n+2):
    sheet.cell(row = 1, column = i).font = fontObj
    sheet.cell(row = 1, column = i).value = i - 1
    sheet.cell(row = i, column=1).font = fontObj
    sheet.cell(row = i, column=1).value = i - 1
    
for i in range(1, n+1):
    for j in range(1, n+1):
        sheet.cell(i+1, j+1).value = sheet.cell(1, j+1).value * sheet.cell(i+1, 1).value

wb.save('multiplicationTables.xlsx')