# -*- coding: utf-8 -*-
"""
Created on Sun Oct 25 01:15:59 2020

@author: Ankit Parashar
"""
#! python3
# sheetInverter.py
import openpyxl

wb = openpyxl.load_workbook('newProduce.xlsx')
sheet = wb.active
sheetData = []
for i in range(sheet.max_column):
    columnData = []
    for j in range(sheet.max_row):
        columnData.append(sheet.cell(row = j+1, column=i+1).value)
    sheetData.append(columnData)
    
newWb = openpyxl.Workbook()
newSheet = newWb.active

for i in range(len(sheetData)):
    for j in range(len(sheetData[i])):
        newSheet.cell(row = i+1, column = j+1).value = sheetData[i][j]
    
newWb.save('newProduce_inverted.xlsx')